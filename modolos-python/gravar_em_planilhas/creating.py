from pathlib import Path

import openpyxl 

from openpyxl.worksheet.worksheet import Worksheet 

ROOT_FOLDER = Path(__file__).parent
WORLBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = openpyxl.Workbook()

sheet_name = 'Minha planilha'

# Criando uma nova pasta de trabalho

workbook.create_sheet(sheet_name, 0)

worksheet: Worksheet = workbook[sheet_name]


# criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')


# Remover uma planilha
workbook.remove(workbook['Sheet'])

studentes = [
    ['joão', 14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz', 15, 8.8],
    ['Alberto', 16, 10]
]

# for i, student_row in enumerate(studentes, start=2):
#     for j, student_colum in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_colum)

for student in studentes:
    worksheet.append(student)


workbook.save(WORLBOOK_PATH)

