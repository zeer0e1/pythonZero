from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet 

ROOT_FOLDER = Path(__file__).parent
WORLBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando arquivo do excel
workbook: Workbook = load_workbook(WORLBOOK_PATH)

# Nome para a planilha
sheet_name = 'Minha planilha'


# Selecionou a planilha
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == 'Maria':
            worksheet.cell(cell.row, 2, 34)
            
    print()

# worksheet['B3'].value = 14

workbook.save(WORLBOOK_PATH)